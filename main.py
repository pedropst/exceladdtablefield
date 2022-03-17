from copy import copy
from pathlib import Path
from tkinter import ttk
from tkinter import *
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import openpyxl
import os
import threading
import sys

def defining_software_path():
    if getattr(sys, 'frozen', False):
        BASE_DIR = os.path.dirname(sys.executable)
    elif __file__:
        BASE_DIR = Path(__file__).resolve().parent
    return BASE_DIR

BASE_DIR = defining_software_path()

def change_cell_type(cell, header):
    if cell.value != None:
        if header == 'Date':
            cell.number_format = 'DD/MM/YYYY'

def auto_size(sheet, cell):
    if (cell.value != None) and (len(cell.value) > sheet.column_dimensions[cell.column_letter].width):
        sheet.column_dimensions[cell.column_letter].width = len(cell.value)

def save_new_file(sheet, path):
    header_cells = []
    header = []
    for column in sheet.iter_cols(min_col=0, max_col=len(sheet.column_dimensions), min_row=1, max_row=1):
        for cell in column:
            if cell.value != None:
                header_cells.append(cell)

    for hcell in header_cells:
        column_values = []
        for rows in sheet.iter_cols(min_col=hcell.column, max_col=hcell.column, min_row=1, max_row=sheet.max_row):
            for c in rows:
                column_values.append(c)
        header.append(column_values)

    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Report"
    last_row = ws.max_row - 1
    for l in header:
        if last_row != 0:
            for c in range(1, len(l)):
                ws.cell(last_row + l[c].row, header.index(l) + 1, l[c].value)
        else:
            for c in range(0, len(l)):
                if c == 0:
                    new_cell = ws.cell(last_row + l[c].row, header.index(l) + 1, l[c].value)
                    new_cell.font = Font(bold=True, size=13)
                    new_cell.alignment = Alignment(horizontal='center')
                else:
                    new_cell = ws.cell(last_row + l[c].row, header.index(l) + 1, l[c].value)
                    auto_size(ws, new_cell)
                    change_cell_type(new_cell, ws.cell(1, new_cell.column).value)
    book.save(path)

def add_data_to_master_file(sheet, path):
    header_cells = []
    header = []
    for column in sheet.iter_cols(min_col=0, max_col=len(sheet.column_dimensions), min_row=1, max_row=1):
        for cell in column:
            if cell.value != None:
                header_cells.append(cell)

    for hcell in header_cells:
        column_values = []
        for rows in sheet.iter_cols(min_col=hcell.column, max_col=hcell.column, min_row=1, max_row=sheet.max_row):
            for c in rows:
                column_values.append(c)
        header.append(column_values)

    if path != '':
        book = openpyxl.load_workbook(path)
        ws = book.active
        last_row = ws.max_row - 1
        for l in header:
            if last_row != 0:
                for c in range(1, len(l)):
                    ws.cell(last_row + l[c].row, header.index(l) + 1, l[c].value)
            else:
                for c in range(0, len(l)):
                    if c == 0:
                        new_cell = ws.cell(last_row + l[c].row, header.index(l) + 1, l[c].value)
                        new_cell.font = Font(bold=True, size=13)
                        new_cell.alignment = Alignment(horizontal='center')
                    else:
                        new_cell = ws.cell(last_row + l[c].row, header.index(l) + 1, l[c].value)
                        auto_size(ws, new_cell)
                        change_cell_type(new_cell, ws.cell(1, new_cell.column).value)
        book.save(path)
    else:
        book = openpyxl.Workbook()
        ws = book.active
        ws.title = "Master File"
        software_path = defining_software_path()
        book.save(os.path.join(software_path, 'master_file.xlsx'))
        add_data_to_master_file(sheet, os.path.join(software_path, 'master_file.xlsx'))
    
def get_category_dictionary(path):
    cat_dic = {}
    if os.path.exists(os.path.join(path, 'category_dictionary.xlsx')):
        category_book = openpyxl.load_workbook(os.path.join(path, 'category_dictionary.xlsx'))
        book = category_book.active

        for rows in book.iter_rows(min_col=1, max_col=1, min_row=1, max_row=book.max_row):
            for cell in rows:
                cat_dic[cell.value] = book.cell(cell.row, 2).value

        cat_dic.pop('MERCHANT')
        return cat_dic
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Category Dictionary"
        new_cell = sheet.cell(1, 1, 'MERCHANT')
        new_cell.font = Font(bold=True, size=13)
        new_cell.alignment = Alignment(horizontal='center')

        new_cell = sheet.cell(1, 2, 'CATEGORY')
        new_cell.font = Font(bold=True, size=13)
        new_cell.alignment = Alignment(horizontal='center')
        
        software_path = defining_software_path()
        book.save(os.path.join(path, 'category_dictionary.xlsx'))
        get_category_dictionary(path)

def add_to_category_dictionary(path, list_to_add):
    category_book = openpyxl.load_workbook(os.path.join(path, 'category_dictionary.xlsx'))
    book = category_book.active

    for k in list_to_add.keys():
        book.cell(book.max_row + 1, 1, k)
        book.cell(book.max_row, 2, list_to_add[k])

    category_book.save(os.path.join(path, 'category_dictionary.xlsx'))

def handle_merged_cells_shifting_and_add_new_columns(sheet, target_column_number):
    select_merged_cells = []
    for merged_cell in sheet.merged_cells.ranges:
        if merged_cell.min_col >= target_column_number:
            select_merged_cells.append(merged_cell)

    for mc in select_merged_cells:
        mc.shift(2,0)

    sheet.insert_cols(target_column_number, 2)

def changing_column_width(sheet):
    for col in sheet.column_dimensions:
        sheet.column_dimensions[col].width = 14.28


class Categories(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        self.input_button = None
        self.choosed = None
        self.path_report = None
        self.path_master = None
        self.software_path = defining_software_path()
        self.main_window = None

    def callback(self):
        pass

    def call(self, path_report, path_master, root):
        self.path_report = path_report
        self.path_master = path_master
        self.main_window = root
        self.start()

    def run(self):
        books = openpyxl.load_workbook(self.path_report)
        target_column_cell = None
        category_column_number = None
        merchant_column_number = None
        sheet = None

        merchant_content = {}

        for sheet in books:
            if sheet.title == 'Report':
                for column in sheet.iter_cols(min_col=0, max_col=len(sheet.column_dimensions), min_row=1, max_row=1):
                    for cell in column:
                        if str(cell.value) == 'Description':
                            target_column_cell = cell
                for rows in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=target_column_cell.column, max_col=target_column_cell.column):
                    for cell in rows:
                        content = cell.value
                        if content != None:
                            if content.__contains__(',USD '):
                                merchant_content[cell] = content.split(',USD ')[1][:-3]
                            elif content.__contains__(',AED '):
                                merchant_content[cell] = content.split(',AED ')[1][:-3]
                            else:
                                merchant_content[cell] = content
                        else:
                            break
                                
                handle_merged_cells_shifting_and_add_new_columns(sheet, target_column_cell.column)

                changing_column_width(sheet)

                category_column_number = target_column_cell.column - 2
                merchant_column_number = target_column_cell.column - 1

                sheet.cell(1, category_column_number, 'Category')
                sheet.cell(1, merchant_column_number, 'Merchant')

                def call(entry, button, label):
                    self.choosed = entry.get()
                    button.destroy()
                    label.destroy()
                    entry.destroy()

                    self.input_button = None

                new_data = {}
                old_data = {}
                
                for k in merchant_content.keys():
                    old_data = get_category_dictionary(BASE_DIR)
                    if old_data == None:
                        old_data = {}
                    sheet.cell(k.row, merchant_column_number, merchant_content[k])
                    if not old_data.keys().__contains__(merchant_content[k]):
                        self.input_button = Tk()
                        self.input_button.geometry('300x150' + "+" + str(self.main_window.winfo_x()) + "+" + str(self.main_window.winfo_y()))
                        self.input_button.title('Add/assign category')
                        l = Label(self.main_window, text=merchant_content[k])
                        l.grid(row=3, column=2, padx=10, pady=10, sticky='W')
                        e = ttk.Combobox(self.main_window, width=20)
                        e['values'] = list(set(old_data.values()))
                        e.grid(row=3, column=1, padx=10, pady=10, sticky='W')
                        b = Button(self.main_window, text='OK', command=lambda: call(e, b, l))
                        b.grid(row=3, column=0, padx=10, pady=10, sticky='W')

                        while self.input_button != None:
                            continue
                        
                        new_data[merchant_content[k]] = self.choosed
                        sheet.cell(k.row, category_column_number, self.choosed)

                        self.choosed = ''
                    else:
                        sheet.cell(k.row, category_column_number, old_data[merchant_content[k]])

                    add_to_category_dictionary(BASE_DIR, new_data)
                    new_data.clear()              

                add_data_to_master_file(sheet, self.path_master)

        report_filename = self.path_report.split('/')[-1]
        report_directory = self.path_report.replace(report_filename, '')

        save_new_file(sheet, os.path.join(report_directory, report_filename.split('.xlsx')[0] + '_with_categories.xlsx'))

